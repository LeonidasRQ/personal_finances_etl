"""
Excel utility functions for handling Excel operations.
"""

import logging
from typing import Any
from copy import copy
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries

logger = logging.getLogger(__name__)


class ExcelStyleCopier:
    """Utility class for copying Excel cell styles."""
    
    def copy_cell_style(self, source_cell: Any, target_cell: Any) -> None:
        """
        Copy all styling from source cell to target cell.
        
        Args:
            source_cell: Source cell to copy style from
            target_cell: Target cell to apply style to
        """
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    
    def copy_row_styles(self, worksheet: Worksheet, source_row: int, target_row: int, 
                       min_col: int, max_col: int) -> None:
        """
        Copy styles from source row to target row.
        
        Args:
            worksheet: Excel worksheet
            source_row: Row number to copy from
            target_row: Row number to copy to
            min_col: Starting column
            max_col: Ending column
        """
        for col in range(min_col, max_col + 1):
            source_cell = worksheet.cell(row=source_row, column=col)
            target_cell = worksheet.cell(row=target_row, column=col)
            self.copy_cell_style(source_cell, target_cell)
    
    def copy_formula_with_row_adjustment(self, source_cell: Any, target_cell: Any, 
                                       source_row: int, target_row: int) -> None:
        """
        Copy formula from source cell to target cell with row number adjustment.
        
        Args:
            source_cell: Source cell with formula
            target_cell: Target cell to receive adjusted formula
            source_row: Source row number
            target_row: Target row number
        """
        if source_cell.data_type == 'f':  # Formula
            formula = source_cell.value
            adjusted_formula = formula.replace(str(source_row), str(target_row))
            target_cell.value = adjusted_formula
        else:
            target_cell.value = source_cell.value


class ValidationExtender:
    """Utility class for extending Excel data validations."""
    
    def extend_validations(self, worksheet: Worksheet, min_row: int, 
                         max_row_before: int, max_row_after: int) -> None:
        """
        Extend data validations to cover new rows.
        
        Args:
            worksheet: Excel worksheet
            min_row: Minimum row in the table
            max_row_before: Maximum row before adding new data
            max_row_after: Maximum row after adding new data
        """
        for dv in worksheet.data_validations.dataValidation:
            new_sqref = []
            for sq in dv.sqref:
                minc, minr, maxc, maxr = range_boundaries(str(sq))
                if min_row <= max_row_before <= maxr:
                    new_range = (f"{worksheet.cell(row=minr, column=minc).coordinate}:"
                               f"{worksheet.cell(row=max_row_after, column=maxc).coordinate}")
                    new_sqref.append(new_range)
                else:
                    new_sqref.append(str(sq))
            dv.sqref = ' '.join(new_sqref)


class ExcelTableManager:
    """Utility class for managing Excel tables."""
    
    def update_table_reference(self, table: Any, min_row: int, min_col: int, 
                             max_row: int, max_col: int, worksheet: Worksheet) -> None:
        """
        Update table reference to include new rows.
        
        Args:
            table: Excel table object
            min_row: Minimum row
            min_col: Minimum column
            max_row: Maximum row
            max_col: Maximum column
            worksheet: Excel worksheet
        """
        start_cell = worksheet.cell(row=min_row, column=min_col).coordinate
        end_cell = worksheet.cell(row=max_row, column=max_col).coordinate
        table.ref = f"{start_cell}:{end_cell}"
        logger.debug(f"Updated table reference to {table.ref}")
