"""
Data loading module for writing to Excel files.
"""

import logging
import sys
from typing import List
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from models.data_models import MappedTransaction
from utils.excel_utils import ExcelStyleCopier, ValidationExtender, ExcelTableManager
from utils.logging_config import get_logger

logger = get_logger('loaders')


class ExcelLoader:
    """Loads transformed data into Excel files."""
    
    def __init__(self, dest_file: str):
        """
        Initialize the Excel loader.
        
        Args:
            dest_file: Path to the destination Excel file
        """
        self.dest_file = Path(dest_file)
        self.style_copier = ExcelStyleCopier()
        self.validation_extender = ValidationExtender()
        self.table_manager = ExcelTableManager()
    
    def load(self, transactions: List[MappedTransaction], dry_run: bool = False) -> int:
        """
        Load transactions into Excel file.
        
        Args:
            transactions: List of mapped transactions to load
            dry_run: If True, don't actually save the file
            
        Returns:
            Number of transactions loaded
            
        Raises:
            FileNotFoundError: If destination file doesn't exist
            Exception: If loading fails
        """
        if not transactions:
            logger.info("No transactions to load")
            return 0
        
        if dry_run:
            logger.info(f"[DRY RUN] Would insert {len(transactions)} rows into {self.dest_file}")
            return len(transactions)
        
        try:
            logger.info(f"Starting load of {len(transactions)} transactions to {self.dest_file}")
            
            # Validate destination file exists
            if not self.dest_file.exists():
                raise FileNotFoundError(f"Destination file not found: {self.dest_file}")
            
            # Load workbook
            wb = load_workbook(str(self.dest_file))
            ws = wb.active
            
            # Get table information
            if not ws.tables:
                raise ValueError("Destination file must contain at least one table")
            
            table = list(ws.tables.values())[0]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            
            logger.debug(f"Table range: {table.ref} (rows {min_row}-{max_row}, cols {min_col}-{max_col})")
            
            # Insert transactions
            original_max_row = max_row
            for transaction in transactions:
                max_row = self._insert_transaction_row(
                    ws, transaction, max_row + 1, min_col, max_col, original_max_row
                )
            
            # Update table and validations
            self._update_table_structure(ws, table, min_row, min_col, max_row, max_col, 
                                       original_max_row, len(transactions))
            
            # Save workbook
            wb.save(str(self.dest_file))
            
            logger.info(f"✅ Successfully loaded {len(transactions)} transactions to {self.dest_file}")
            return len(transactions)
            
        except Exception as e:
            logger.error(f"Failed to load transactions to {self.dest_file}: {e}")
            raise
    
    def _insert_transaction_row(self, ws, transaction: MappedTransaction, row: int, 
                              min_col: int, max_col: int, template_row: int) -> int:
        """
        Insert a single transaction row with proper formatting.
        
        Args:
            ws: Worksheet object
            transaction: Transaction to insert
            row: Row number to insert at
            min_col: Minimum column of table
            max_col: Maximum column of table
            template_row: Row to copy styles from
            
        Returns:
            Row number of inserted row
        """
        # Prepare row data
        expense = transaction.amount if transaction.flow_type == "Egreso" else ""
        income = transaction.amount if transaction.flow_type == "Ingreso" else ""
        
        row_data = [
            transaction.date,
            transaction.flow_type,
            transaction.category,
            transaction.subcategory,
            transaction.description,
            expense,
            income
        ]
        
        # Insert values
        for idx, value in enumerate(row_data, start=min_col):
            if idx <= max_col:  # Don't exceed table columns
                ws.cell(row=row, column=idx, value=value)
        
        # Copy styles from template row
        self.style_copier.copy_row_styles(ws, template_row, row, min_col, max_col)
        
        # Handle balance column (last column) with formula
        if max_col >= min_col + len(row_data):
            balance_col = max_col
            source_balance_cell = ws.cell(row=template_row, column=balance_col)
            target_balance_cell = ws.cell(row=row, column=balance_col)
            
            self.style_copier.copy_formula_with_row_adjustment(
                source_balance_cell, target_balance_cell, template_row, row
            )
            self.style_copier.copy_cell_style(source_balance_cell, target_balance_cell)
        
        logger.debug(f"Inserted transaction at row {row}: {transaction.description}")
        return row
    
    def _update_table_structure(self, ws, table, min_row: int, min_col: int, 
                              max_row: int, max_col: int, original_max_row: int, 
                              rows_added: int) -> None:
        """
        Update table reference and data validations after adding rows.
        
        Args:
            ws: Worksheet object
            table: Table object
            min_row: Minimum row of table
            min_col: Minimum column of table
            max_row: New maximum row of table
            max_col: Maximum column of table
            original_max_row: Original maximum row before adding data
            rows_added: Number of rows added
        """
        # Extend data validations
        self.validation_extender.extend_validations(
            ws, min_row, original_max_row, max_row
        )
        
        # Update table reference
        self.table_manager.update_table_reference(
            table, min_row, min_col, max_row, max_col, ws
        )
        
        logger.debug(f"Updated table structure: added {rows_added} rows")
    
    def validate_destination_format(self) -> bool:
        """
        Validate that the destination file has the expected format.
        
        Returns:
            True if format is valid, False otherwise
        """
        try:
            if not self.dest_file.exists():
                logger.error(f"Destination file does not exist: {self.dest_file}")
                return False
            
            wb = load_workbook(str(self.dest_file))
            ws = wb.active
            
            # Check for tables
            if not ws.tables:
                logger.error("Destination file must contain at least one table")
                return False
            
            # Check table structure
            table = list(ws.tables.values())[0]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            
            # Expect at least 7 columns for the financial data format
            expected_columns = 7
            actual_columns = max_col - min_col + 1
            
            if actual_columns < expected_columns:
                logger.error(
                    f"Table has insufficient columns: {actual_columns} < {expected_columns}"
                )
                return False
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to validate destination format: {e}")
            return False
