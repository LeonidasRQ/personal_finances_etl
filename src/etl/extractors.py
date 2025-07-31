"""
Data extraction module for reading from Excel files.
"""

import logging
import sys
from typing import List
from openpyxl import load_workbook
from pathlib import Path

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from models.data_models import RawTransaction
from utils.date_utils import parse_date
from utils.logging_config import get_logger

logger = get_logger('extractors')


class ExcelExtractor:
    """Extracts financial data from Excel files."""
    
    def __init__(self, file_path: str):
        """
        Initialize the Excel extractor.
        
        Args:
            file_path: Path to the Excel file to extract from
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    def extract(self) -> List[RawTransaction]:
        """
        Extract data from Excel file and return structured data.
        
        Returns:
            List of RawTransaction objects
            
        Raises:
            Exception: If extraction fails
        """
        try:
            logger.info(f"Starting extraction from {self.file_path}")
            
            wb = load_workbook(str(self.file_path), data_only=True)
            ws = wb.active
            
            transactions = []
            skipped_rows = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extract data from row (first 6 columns)
                    date, description, value, _, _, category_origin = row[:6]
                    
                    # Skip rows with missing essential data
                    if date is None or value is None:
                        skipped_rows += 1
                        logger.debug(f"Skipping row {row_num}: missing date or value")
                        continue
                    
                    # Parse and validate data
                    parsed_date = parse_date(date)
                    clean_value = abs(float(value))
                    clean_description = str(description) if description else ""
                    clean_category = str(category_origin) if category_origin else "Others"
                    
                    # Create transaction object
                    transaction = RawTransaction(
                        date=parsed_date,
                        description=clean_description,
                        value=clean_value,
                        category_origin=clean_category
                    )
                    
                    transactions.append(transaction)
                    
                except Exception as e:
                    logger.warning(f"Error processing row {row_num}: {e}")
                    skipped_rows += 1
                    continue
            
            logger.info(
                f"Extraction completed: {len(transactions)} transactions extracted, "
                f"{skipped_rows} rows skipped"
            )
            
            return transactions
            
        except Exception as e:
            logger.error(f"Failed to extract data from {self.file_path}: {e}")
            raise
    
    def validate_file_format(self) -> bool:
        """
        Validate that the Excel file has the expected format.
        
        Returns:
            True if format is valid, False otherwise
        """
        try:
            wb = load_workbook(str(self.file_path), data_only=True)
            ws = wb.active
            
            # Check if file has enough columns
            if ws.max_column < 6:
                logger.error(f"Excel file has insufficient columns: {ws.max_column} < 6")
                return False
            
            # Check if file has data rows
            if ws.max_row < 2:
                logger.error("Excel file has no data rows")
                return False
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to validate file format: {e}")
            return False
