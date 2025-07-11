import datetime
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from copy import copy

# ----------------------
# CONFIG
# ----------------------
excel_file = "copia_finanzas.xlsx"
sheet_name = "cashflow"
table_name = "flujo_dinero"

source_file = "mobills_transactions.xlsx"
source_sheet = "Receitas e Despesas"

# ----------------------
# CATEGORÍAS Y SUBCATEGORÍAS
# ----------------------
category_to_subcategory = {
    # Necesidades
    "Apartament Expenses": "Servicios Apartamento",
    "Corozal": "Corozal",
    "Health": "Salud",
    "Education": "Educación",
    "Electronics": "Celular",
    "Pets": "Mascotas",
    "Transportation": "Transporte",
    "Supermarket": "Mercado",
    "Mami": "Mami",
    "Personal Care": "Cuidado Personal",

    # Gustos
    "Entertainment": "Entretenimiento",
    "Friends": "Amigos",
    "Gifts": "Regalos",
    "Adjustment*": "Random",
    "Random": "Random",
    "Restaurant": "Restaurantes",
    "Loans": "Prestamos",

    # Ahorros_y_Deudas
    "CC Interests": "Deudas",
    "Debt": "Deudas",

    # Ingresos
    "Salary": "Salario",
    "Siblings Support": "Ayuda",
    "Others": "Otro"

    # Default
    # Si hay algún otro lo marcaría como "Otro"
}

subcategory_to_category = {
    # Necesidades
    "Servicios Apartamento": "Necesidades",
    "Corozal": "Necesidades",
    "Salud": "Necesidades",
    "Educación": "Necesidades",
    "Mascotas": "Necesidades",
    "Transporte": "Necesidades",
    "Mercado": "Necesidades",
    "Mami": "Necesidades",
    "Cuidado Personal": "Necesidades",
    "Celular": "Necesidades",

    # Gustos
    "Entretenimiento": "Gustos",
    "Amigos": "Gustos",
    "Regalos": "Gustos",
    "Random": "Gustos",
    "Restaurantes": "Gustos",
    "Prestamos": "Gustos",

    # Ahorros_y_Deudas
    "Deudas": "Ahorros_y_Deudas",

    # Ingresos
    "Salario": "Ingreso",
    "Ayuda": "Ingreso",
    "Guardias": "Ingreso",
    "Arriendo": "Ingreso",
    "Prestamos": "Ingreso",
    "Prima": "Ingreso",
    "Otro": "Ingreso"  # fallback para ingresos si aplica
}

# ----------------------
# CARGAR EXCEL ORIGEN
# ----------------------
source_wb = load_workbook(source_file)
source_ws = source_wb[source_sheet]

entries = []
for row in source_ws.iter_rows(min_row=2, values_only=True):
    date, description, value, category_origin = row[0], row[1], row[2], row[5]

    if date is None or value is None:
        continue

    subcategoria = category_to_subcategory.get(category_origin, "Otro")
    if subcategoria == "Otro":
        print(f"⚠️ WARNING: categoría origen '{category_origin}' no reconocida, se asigna 'Otro'")

    categoria_base = subcategory_to_category.get(subcategoria, "Ingreso")
    tipo_flujo = "Ingreso" if categoria_base == "Ingreso" else "Egreso"
    monto = abs(value)

    if tipo_flujo == "Ingreso":
        categoria = subcategoria
        subcategoria_final = ""
    else:
        categoria = categoria_base
        subcategoria_final = subcategoria

    entries.append({
        "fecha": date.strftime("%Y-%m-%d") if isinstance(date, datetime.datetime) else date,
        "tipo_flujo": tipo_flujo,
        "categoria": categoria,
        "subcategoria": subcategoria_final,
        "descripcion": description or "",
        "monto": monto
    })
    
    # ORDENA LAS ENTRIES POR FECHA 
    entries.sort(key=lambda x: x["fecha"])

print(f"✅ {len(entries)} registros cargados y categorizados desde {source_file}")

# ----------------------
# CARGAR EXCEL DESTINO
# ----------------------
wb = load_workbook(excel_file)
ws = wb[sheet_name]

table = ws.tables[table_name]
min_col, min_row, max_col, max_row = range_boundaries(table.ref)

# ----------------------
# FUNCIÓN PARA EXTENDER VALIDACIONES
# ----------------------
def extend_validations(ws, max_row_before, max_row_after):
    for dv in ws.data_validations.dataValidation:
        new_sqref = []
        for sq in dv.sqref:
            minc, minr, maxc, maxr = range_boundaries(str(sq))
            if min_row <= max_row_before <= maxr:
                new_sqref.append(f"{ws.cell(row=minr, column=minc).coordinate}:{ws.cell(row=max_row_after, column=maxc).coordinate}")
            else:
                new_sqref.append(str(sq))
        dv.sqref = ' '.join(new_sqref)

# ----------------------
# INSERTAR CADA REGISTRO
# ----------------------
for entry in entries:
    tipo_flujo = entry["tipo_flujo"]
    monto = entry["monto"]

    # Convertir fecha string a datetime.date
    fecha_str = entry["fecha"]
    fecha = datetime.datetime.strptime(fecha_str, "%d/%m/%Y").date()

    if tipo_flujo == "Ingreso":
        egreso = ''
        ingreso = monto
    elif tipo_flujo == "Egreso":
        egreso = monto
        ingreso = ''
    else:
        raise ValueError(f"⚠ Tipo de Flujo inválido: {tipo_flujo}")

    new_row_data = [
        fecha,
        tipo_flujo,
        entry["categoria"],
        entry["subcategoria"],
        entry["descripcion"],
        egreso,
        ingreso
    ]

    next_row = max_row + 1

    # Insertar valores
    for idx, value in enumerate(new_row_data, start=min_col):
        ws.cell(row=next_row, column=idx, value=value)

    # Copiar estilos
    for idx in range(min_col, max_col + 1):
        source_cell = ws.cell(row=max_row, column=idx)
        target_cell = ws.cell(row=next_row, column=idx)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    # Copiar fórmula balance
    balance_col = max_col
    source_balance_cell = ws.cell(row=max_row, column=balance_col)
    target_balance_cell = ws.cell(row=next_row, column=balance_col)

    if source_balance_cell.data_type == 'f':
        target_balance_cell.value = source_balance_cell.value.replace(str(max_row), str(next_row))
    else:
        target_balance_cell.value = source_balance_cell.value

    target_balance_cell.font = copy(source_balance_cell.font)
    target_balance_cell.border = copy(source_balance_cell.border)
    target_balance_cell.fill = copy(source_balance_cell.fill)
    target_balance_cell.number_format = copy(source_balance_cell.number_format)
    target_balance_cell.protection = copy(source_balance_cell.protection)
    target_balance_cell.alignment = copy(source_balance_cell.alignment)

    # Actualizar fila actual
    max_row = next_row

# ----------------------
# EXTENDER VALIDACIONES Y TABLA
# ----------------------
extend_validations(ws, max_row - len(entries), max_row)
table.ref = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"

# ----------------------
# GUARDAR
# ----------------------
wb.save(excel_file)
print(f"✅ {len(entries)} filas agregadas con estilos, fórmulas y validaciones.")
