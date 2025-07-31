import unittest
import datetime
import json
import tempfile
import os
from unittest.mock import Mock, patch, mock_open, MagicMock
from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.datavalidation import DataValidation

import etl_finance


class TestETLFinance(unittest.TestCase):
    
    def setUp(self):
        """Set up test fixtures before each test method."""
        self.sample_mapping = {
            "category_to_subcategory": {
                "Salary": "Salario",
                "Supermarket": "Mercado",
                "Transportation": "Transporte",
                "Loans": "Prestamos",
                "Others": "Otro"
            },
            "subcategory_to_category": {
                "Salario": "Ingreso",
                "Mercado": "Necesidades",
                "Transporte": "Necesidades",
                "Prestamos": "Gustos",
                "Otro": "Ingreso"
            }
        }
        
        self.sample_raw_data = [
            {
                "date": datetime.date(2025, 1, 15),
                "description": "Monthly salary",
                "value": 5000.0,
                "category_origin": "Salary"
            },
            {
                "date": datetime.date(2025, 1, 16),
                "description": "Grocery shopping",
                "value": 150.0,
                "category_origin": "Supermarket"
            },
            {
                "date": datetime.date(2025, 1, 17),
                "description": "Bus fare",
                "value": 25.0,
                "category_origin": "Transportation"
            }
        ]


class TestLoadMapping(TestETLFinance):
    
    def test_load_mapping_success(self):
        """Test successful loading of mapping file."""
        mock_file_content = json.dumps(self.sample_mapping)
        
        with patch("builtins.open", mock_open(read_data=mock_file_content)):
            result = etl_finance.load_mapping("dummy_path.json")
            
        self.assertEqual(result, self.sample_mapping)
    
    def test_load_mapping_file_not_found(self):
        """Test handling of missing mapping file."""
        with patch("builtins.open", side_effect=FileNotFoundError):
            with self.assertRaises(FileNotFoundError):
                etl_finance.load_mapping("nonexistent.json")
    
    def test_load_mapping_invalid_json(self):
        """Test handling of invalid JSON in mapping file."""
        invalid_json = "{ invalid json content"
        
        with patch("builtins.open", mock_open(read_data=invalid_json)):
            with self.assertRaises(json.JSONDecodeError):
                etl_finance.load_mapping("invalid.json")


class TestLoadSourceData(TestETLFinance):
    
    @patch('etl_finance.load_workbook')
    def test_load_source_data_success(self, mock_load_workbook):
        """Test successful loading of source data from Excel."""
        # Mock workbook and worksheet
        mock_wb = Mock()
        mock_ws = Mock()
        mock_wb.active = mock_ws
        mock_load_workbook.return_value = mock_wb
        
        # Mock row data
        mock_rows = [
            (datetime.date(2025, 1, 15), "Monthly salary", 5000.0, None, None, "Salary"),
            (datetime.date(2025, 1, 16), "Grocery shopping", -150.0, None, None, "Supermarket"),
            (None, "Invalid row", 100.0, None, None, "Others"),  # Should be skipped
            (datetime.date(2025, 1, 17), "Bus fare", None, None, None, "Transportation"),  # Should be skipped
        ]
        mock_ws.iter_rows.return_value = mock_rows
        
        result = etl_finance.load_source_data("dummy_file.xlsx")
        
        expected = [
            {
                "date": datetime.date(2025, 1, 15),
                "description": "Monthly salary",
                "value": 5000.0,
                "category_origin": "Salary"
            },
            {
                "date": datetime.date(2025, 1, 16),
                "description": "Grocery shopping",
                "value": 150.0,  # Should be abs value
                "category_origin": "Supermarket"
            }
        ]
        
        self.assertEqual(result, expected)
        mock_load_workbook.assert_called_once_with("dummy_file.xlsx", data_only=True)
        mock_ws.iter_rows.assert_called_once_with(min_row=2, values_only=True)
    
    @patch('etl_finance.load_workbook')
    def test_load_source_data_empty_description(self, mock_load_workbook):
        """Test handling of empty description in source data."""
        mock_wb = Mock()
        mock_ws = Mock()
        mock_wb.active = mock_ws
        mock_load_workbook.return_value = mock_wb
        
        mock_rows = [
            (datetime.date(2025, 1, 15), None, 5000.0, None, None, "Salary"),
        ]
        mock_ws.iter_rows.return_value = mock_rows
        
        result = etl_finance.load_source_data("dummy_file.xlsx")
        
        self.assertEqual(result[0]["description"], "")


class TestParseDateFunction(TestETLFinance):
    
    def test_parse_date_with_date_object(self):
        """Test parsing when input is already a date object."""
        input_date = datetime.date(2025, 1, 15)
        result = etl_finance.parse_date(input_date)
        self.assertEqual(result, input_date)
    
    def test_parse_date_with_string(self):
        """Test parsing date from string format."""
        input_string = "15/01/2025"
        expected = datetime.date(2025, 1, 15)
        result = etl_finance.parse_date(input_string)
        self.assertEqual(result, expected)
    
    def test_parse_date_invalid_format(self):
        """Test handling of invalid date format."""
        invalid_date = "2025-01-15"  # Wrong format
        with self.assertRaises(ValueError):
            etl_finance.parse_date(invalid_date)


class TestMapData(TestETLFinance):
    
    def test_map_data_income(self):
        """Test mapping of income data."""
        raw_data = [
            {
                "date": "15/01/2025",
                "description": "Monthly salary",
                "value": 5000.0,
                "category_origin": "Salary"
            }
        ]
        
        result = etl_finance.map_data(raw_data, self.sample_mapping)
        
        expected = [
            {
                "date": datetime.date(2025, 1, 15),
                "flow_type": "Ingreso",
                "category": "Salario",
                "subcategory": "",
                "description": "Monthly salary",
                "amount": 5000.0
            }
        ]
        
        self.assertEqual(result, expected)
    
    def test_map_data_expense(self):
        """Test mapping of expense data."""
        raw_data = [
            {
                "date": "16/01/2025",
                "description": "Grocery shopping",
                "value": 150.0,
                "category_origin": "Supermarket"
            }
        ]
        
        result = etl_finance.map_data(raw_data, self.sample_mapping)
        
        expected = [
            {
                "date": datetime.date(2025, 1, 16),
                "flow_type": "Egreso",
                "category": "Necesidades",
                "subcategory": "Mercado",
                "description": "Grocery shopping",
                "amount": 150.0
            }
        ]
        
        self.assertEqual(result, expected)
    
    def test_map_data_unknown_category(self):
        """Test mapping with unknown category defaults to 'Otro'."""
        raw_data = [
            {
                "date": "17/01/2025",
                "description": "Unknown expense",
                "value": 100.0,
                "category_origin": "Unknown Category"
            }
        ]
        
        result = etl_finance.map_data(raw_data, self.sample_mapping)
        
        expected = [
            {
                "date": datetime.date(2025, 1, 17),
                "flow_type": "Ingreso",  # "Otro" maps to "Ingreso"
                "category": "Otro",
                "subcategory": "",
                "description": "Unknown expense",
                "amount": 100.0
            }
        ]
        
        self.assertEqual(result, expected)
    
    def test_map_data_sorting(self):
        """Test that mapped data is sorted by date."""
        raw_data = [
            {
                "date": "20/01/2025",
                "description": "Later date",
                "value": 100.0,
                "category_origin": "Salary"
            },
            {
                "date": "15/01/2025",
                "description": "Earlier date",
                "value": 200.0,
                "category_origin": "Salary"
            }
        ]
        
        result = etl_finance.map_data(raw_data, self.sample_mapping)
        
        # Should be sorted by date ascending
        self.assertEqual(result[0]["date"], datetime.date(2025, 1, 15))
        self.assertEqual(result[1]["date"], datetime.date(2025, 1, 20))


class TestExtendValidations(TestETLFinance):
    
    def test_extend_validations(self):
        """Test extending data validations when new rows are added."""
        # Create a mock worksheet with data validation
        mock_ws = Mock()
        mock_dv = Mock()
        mock_dv.sqref = ["A1:A10"]
        mock_ws.data_validations.dataValidation = [mock_dv]
        
        # Mock cells for coordinate generation
        mock_ws.cell.side_effect = lambda row, column: Mock(coordinate=f"{chr(64+column)}{row}")
        
        with patch('etl_finance.range_boundaries', return_value=(1, 1, 1, 10)):
            etl_finance.extend_validations(mock_ws, 2, 10, 15)
        
        # Check that sqref was updated
        self.assertEqual(mock_dv.sqref, "A1:A15")


class TestInsertIntoExcel(TestETLFinance):
    
    @patch('etl_finance.load_workbook')
    @patch('etl_finance.range_boundaries')
    @patch('etl_finance.extend_validations')
    def test_insert_into_excel_basic(self, mock_extend_validations, mock_range_boundaries, mock_load_workbook):
        """Test basic insertion of data into Excel."""
        # Setup mocks
        mock_wb = Mock()
        mock_ws = Mock()
        mock_table = Mock()
        mock_wb.active = mock_ws
        mock_ws.tables.values.return_value = [mock_table]
        mock_load_workbook.return_value = mock_wb
        mock_range_boundaries.return_value = (1, 2, 8, 10)  # min_col, min_row, max_col, max_row
        
        # Mock cells
        cell_dict = {}
        def mock_cell(row, column, value=None):
            key = (row, column)
            if key not in cell_dict:
                cell_dict[key] = Mock()
                cell_dict[key].coordinate = f"{chr(64+column)}{row}"
                cell_dict[key].has_style = True
                cell_dict[key].data_type = 'n'  # numeric
                cell_dict[key].value = 100.0
            if value is not None:
                cell_dict[key].value = value
            return cell_dict[key]
        
        mock_ws.cell = mock_cell
        
        # Test data
        entries = [
            {
                "date": datetime.date(2025, 1, 15),
                "flow_type": "Egreso",
                "category": "Necesidades",
                "subcategory": "Mercado",
                "description": "Grocery shopping",
                "amount": 150.0
            }
        ]
        
        result = etl_finance.insert_into_excel(entries, "dummy_file.xlsx")
        
        # Verify workbook was loaded
        mock_load_workbook.assert_called_once_with("dummy_file.xlsx")
        
        # Verify extend_validations was called
        mock_extend_validations.assert_called_once()
        
        # Verify table reference was updated
        self.assertEqual(mock_table.ref, "A2:H11")
        
        # Check return values
        wb, ws, table, max_row = result
        self.assertEqual(wb, mock_wb)
        self.assertEqual(ws, mock_ws)
        self.assertEqual(table, mock_table)
        self.assertEqual(max_row, 11)


class TestMainFunction(TestETLFinance):
    
    @patch('etl_finance.insert_into_excel')
    @patch('etl_finance.map_data')
    @patch('etl_finance.load_source_data')
    @patch('etl_finance.load_mapping')
    def test_main_dry_run(self, mock_load_mapping, mock_load_source_data, 
                         mock_map_data, mock_insert_into_excel):
        """Test main function with dry run option."""
        # Setup mocks
        mock_load_mapping.return_value = self.sample_mapping
        mock_load_source_data.return_value = self.sample_raw_data
        mock_map_data.return_value = [{"test": "data"}]
        
        mock_wb = Mock()
        mock_insert_into_excel.return_value = (mock_wb, Mock(), Mock(), 10)
        
        # Test dry run
        with patch('sys.argv', ['etl_finance.py', '--dry-run']):
            etl_finance.main()
        
        # Verify functions were called
        mock_load_mapping.assert_called_once_with('config/category_mapping.json')
        mock_load_source_data.assert_called_once_with('C:/Users/Leonidas/Downloads/current_mobills_transactions.xlsx')
        mock_map_data.assert_called_once_with(self.sample_raw_data, self.sample_mapping)
        mock_insert_into_excel.assert_called_once_with([{"test": "data"}], 'C:/Users/Leonidas/OneDrive - cecar.edu.co/finanzas_2025.xlsx')
        
        # Verify save was NOT called in dry run
        mock_wb.save.assert_not_called()
    
    @patch('etl_finance.insert_into_excel')
    @patch('etl_finance.map_data')
    @patch('etl_finance.load_source_data')
    @patch('etl_finance.load_mapping')
    def test_main_normal_run(self, mock_load_mapping, mock_load_source_data, 
                            mock_map_data, mock_insert_into_excel):
        """Test main function with normal execution."""
        # Setup mocks
        mock_load_mapping.return_value = self.sample_mapping
        mock_load_source_data.return_value = self.sample_raw_data
        mock_map_data.return_value = [{"test": "data"}]
        
        mock_wb = Mock()
        mock_insert_into_excel.return_value = (mock_wb, Mock(), Mock(), 10)
        
        # Test normal run
        with patch('sys.argv', ['etl_finance.py']):
            etl_finance.main()
        
        # Verify save WAS called in normal run
        mock_wb.save.assert_called_once_with('C:/Users/Leonidas/OneDrive - cecar.edu.co/finanzas_2025.xlsx')


class TestIntegration(TestETLFinance):
    """Integration tests for the complete ETL process."""
    
    def test_end_to_end_mapping_flow(self):
        """Test the complete mapping flow from raw data to final format."""
        # Test data that includes the duplicate "Prestamos" issue
        raw_data = [
            {
                "date": "15/01/2025",
                "description": "Salary payment",
                "value": 5000.0,
                "category_origin": "Salary"
            },
            {
                "date": "16/01/2025",
                "description": "Loan payment made",
                "value": 500.0,
                "category_origin": "Loans"  # This should map to "Prestamos" -> "Gustos"
            }
        ]
        
        result = etl_finance.map_data(raw_data, self.sample_mapping)
        
        # Verify income mapping
        income_entry = next(item for item in result if item["flow_type"] == "Ingreso")
        self.assertEqual(income_entry["category"], "Salario")
        self.assertEqual(income_entry["subcategory"], "")
        
        # Verify expense mapping (loan payment)
        expense_entry = next(item for item in result if item["flow_type"] == "Egreso")
        self.assertEqual(expense_entry["category"], "Gustos")
        self.assertEqual(expense_entry["subcategory"], "Prestamos")


if __name__ == '__main__':
    # Configure logging for tests
    import logging
    logging.disable(logging.CRITICAL)
    
    unittest.main()
