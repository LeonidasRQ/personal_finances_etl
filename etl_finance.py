import argparse
import datetime
import json
import logging
import os

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from copy import copy
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Get environment config
DEST_FILE = os.getenv("DEST_FILE")
SOURCE_FILE = os.getenv("SOURCE_FILE")
MAPPING_FILE = os.getenv("MAPPING_FILE")
LOG_FILE = os.getenv("LOG_FILE")

# Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def load_mapping(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_source_data(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date, description, value, _, _, category_origin = row[:6]
        if date is None or value is None:
            continue
        data.append({
            "date": date,
            "description": description or "",
            "value": abs(value),
            "category_origin": category_origin
        })
    return data


def parse_date(date_str):
    if isinstance(date_str, datetime.date):
        return date_str
    return datetime.datetime.strptime(date_str, "%d/%m/%Y").date()


def map_data(raw_data, mapping):
    category_to_subcategory = mapping["category_to_subcategory"]
    subcategory_to_category = mapping["subcategory_to_category"]

    mapped = []
    for item in raw_data:
        subcategory = category_to_subcategory.get(item["category_origin"], "Otro")
        category_base = subcategory_to_category.get(subcategory, "Ingreso")
        flow_type = "Ingreso" if category_base == "Ingreso" else "Egreso"

        if flow_type == "Ingreso":
            category = subcategory
            final_subcategory = ""
        else:
            category = category_base
            final_subcategory = subcategory

        mapped.append({
            "date": item["date"],
            "flow_type": flow_type,
            "category": category,
            "subcategory": final_subcategory,
            "description": item["description"],
            "amount": item["value"]
        })

    for item in mapped:
        item["date"] = parse_date(item["date"])

    mapped.sort(key=lambda x: x["date"])
    # Sort ascending by date
    mapped.sort(key=lambda x: x["date"])
    return mapped


def extend_validations(ws, min_row, max_row_before, max_row_after):
    for dv in ws.data_validations.dataValidation:
        new_sqref = []
        for sq in dv.sqref:
            minc, minr, maxc, maxr = range_boundaries(str(sq))
            if min_row <= max_row_before <= maxr:
                new_sqref.append(f"{ws.cell(row=minr, column=minc).coordinate}:{ws.cell(row=max_row_after, column=maxc).coordinate}")
            else:
                new_sqref.append(str(sq))
        dv.sqref = ' '.join(new_sqref)


def insert_into_excel(entries, dest_file):
    wb = load_workbook(dest_file)
    ws = wb.active

    # Get first table
    table = list(ws.tables.values())[0]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    for entry in entries:
        flow_type = entry["flow_type"]
        amount = entry["amount"]

        # Fix date
        date = entry["date"]
        if isinstance(date, str):
            try:
                date = datetime.datetime.strptime(date, "%d/%m/%Y").date()
            except ValueError:
                date = datetime.datetime.strptime(date, "%Y/%m/%d").date()

        expense = amount if flow_type == "Egreso" else ""
        income = amount if flow_type == "Ingreso" else ""

        new_row_data = [
            date,
            flow_type,
            entry["category"],
            entry["subcategory"],
            entry["description"],
            expense,
            income
        ]

        next_row = max_row + 1

        # Insert values
        for idx, value in enumerate(new_row_data, start=min_col):
            ws.cell(row=next_row, column=idx, value=value)

        # Copy styles
        for idx in range(min_col, max_col + 1):
            source_cell = ws.cell(row=max_row, column=idx)
            target_cell = ws.cell(row=next_row, column=idx)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        # Copy formula
        balance_col = max_col
        source_balance_cell = ws.cell(row=max_row, column=balance_col)
        target_balance_cell = ws.cell(row=next_row, column=balance_col)
        if source_balance_cell.data_type == 'f':
            target_balance_cell.value = source_balance_cell.value.replace(str(max_row), str(next_row))
        else:
            target_balance_cell.value = source_balance_cell.value

        # Copy style
        target_balance_cell.font = copy(source_balance_cell.font)
        target_balance_cell.border = copy(source_balance_cell.border)
        target_balance_cell.fill = copy(source_balance_cell.fill)
        target_balance_cell.number_format = copy(source_balance_cell.number_format)
        target_balance_cell.protection = copy(source_balance_cell.protection)
        target_balance_cell.alignment = copy(source_balance_cell.alignment)

        max_row = next_row

    # Extend validations & table
    extend_validations(ws, min_row, max_row - len(entries), max_row)
    table.ref = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"

    return wb, ws, table, max_row


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Run without saving changes to the file")
    args = parser.parse_args()

    mapping = load_mapping(MAPPING_FILE)
    raw_data = load_source_data(SOURCE_FILE)
    entries = map_data(raw_data, mapping)
    logger.info(f"Loaded and mapped {len(entries)} records from {SOURCE_FILE}")

    wb, ws, table, last_row = insert_into_excel(entries, DEST_FILE)

    if args.dry_run:
        logger.info(f"[DRY RUN] Would insert {len(entries)} rows into {DEST_FILE}")
    else:
        wb.save(DEST_FILE)
        logger.info(f"✅ Inserted {len(entries)} rows into {DEST_FILE}")


if __name__ == "__main__":
    main()
