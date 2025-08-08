#!/usr/bin/env python3
"""
Diagnostic script to find what changed in your working environment.
"""

import os
import sys
from pathlib import Path

def diagnose_issue():
    print("🔍 Diagnosing ETL Pipeline Issue")
    print("=" * 50)
    
    # Check current directory
    print(f"📍 Current directory: {os.getcwd()}")
    
    # Check Python environment
    print(f"🐍 Python executable: {sys.executable}")
    print(f"🐍 Python version: {sys.version}")
    
    # Check if in virtual environment
    in_venv = hasattr(sys, 'real_prefix') or sys.base_prefix != sys.prefix
    print(f"🔧 In virtual environment: {in_venv}")
    
    # Check file paths from .env
    try:
        from dotenv import load_dotenv
        load_dotenv()
        
        dest_file = os.getenv('DEST_FILE', '').strip('"')
        source_file = os.getenv('SOURCE_FILE', '').strip('"')
        
        print(f"\n📁 File Paths:")
        print(f"   Destination: {dest_file}")
        print(f"   Source: {source_file}")
        
        # Check if files exist
        dest_exists = Path(dest_file).exists() if dest_file else False
        source_exists = Path(source_file).exists() if source_file else False
        
        print(f"   Destination exists: {'✅' if dest_exists else '❌'}")
        print(f"   Source exists: {'✅' if source_exists else '❌'}")
        
        # Check file permissions and access
        if dest_exists:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(dest_file)
                print(f"   ✅ Can open destination file")
                print(f"   📊 Sheets: {wb.sheetnames}")
                print(f"   📋 Active sheet: {wb.active.title}")
                
                # Check for tables
                tables_info = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if ws.tables:
                        tables_info.append(f"{sheet_name}: {list(ws.tables.keys())}")
                
                if tables_info:
                    print(f"   📊 Tables found:")
                    for info in tables_info:
                        print(f"      {info}")
                else:
                    print(f"   ⚠️  No tables found in any sheet")
                
            except PermissionError:
                print(f"   ❌ Permission denied - file might be open in Excel")
            except Exception as e:
                print(f"   ❌ Error opening file: {e}")
        
    except ImportError:
        print("❌ Cannot import dotenv - check if installed")
    except Exception as e:
        print(f"❌ Error loading .env: {e}")
    
    # Check dependencies
    print(f"\n📦 Dependencies:")
    required_packages = ['openpyxl', 'python-dotenv', 'pandas']
    
    for package in required_packages:
        try:
            __import__(package)
            print(f"   ✅ {package}")
        except ImportError:
            print(f"   ❌ {package} - not installed")
    
    # Check if Excel is running
    print(f"\n🔍 Process Check:")
    try:
        import psutil
        excel_processes = [p for p in psutil.process_iter(['name']) if 'excel' in p.info['name'].lower()]
        if excel_processes:
            print(f"   ⚠️  Excel is running - close it and try again")
        else:
            print(f"   ✅ Excel is not running")
    except ImportError:
        print(f"   ⚠️  Cannot check processes (psutil not installed)")

if __name__ == "__main__":
    diagnose_issue()